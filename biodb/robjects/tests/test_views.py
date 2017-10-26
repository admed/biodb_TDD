import PyPDF2
from unit_tests.base import FunctionalTest
from robjects.models import Robject, Name, Tag
from projects.models import Project
from samples.models import Sample
from django.contrib.auth.models import User
from django.core.urlresolvers import reverse
from django_addanother.widgets import AddAnotherWidgetWrapper
from django import forms
from django_addanother.views import CreatePopupMixin
from django.views import generic
from io import BytesIO
from openpyxl import load_workbook
from robjects.views import NameCreateView, TagCreateView
from biodb import settings
from guardian.shortcuts import assign_perm
from tools.history import CustomHistory
from io import BytesIO
from openpyxl import load_workbook
from unittest.mock import patch, call
import datetime


class Robjects_export_to_excel_view_test(FunctionalTest):
    def test_view_returns_404_when_slug_not_match(self):
        self.not_matching_url_slug_helper(self.ROBJECT_EXCEL_URL)

    def test_annonymous_request(self):
        self.annonymous_testing_helper(self.ROBJECT_EXCEL_URL)

    def test_no_visit_permission_request(self):
        self.permission_testing_helper(
            self.ROBJECT_EXCEL_URL,
            error_message="User doesn't have permission: can visit project")

    def test_excel_filename(self):
        user, proj = self.default_set_up_for_robjects_pages()
        r = Robject.objects.create(project=proj, name="robject_1")
        response = self.client.get(f"/projects/{proj.name}/robjects/excel-raport/",
                                   {"robject_1": r.id})
        self.assertEqual(response.get('Content-Disposition'),
                         "attachment; filename=report.xlsx")

    def get_cells_values_from_excel_row(self, response, row_nr):
        with BytesIO(response.content) as f:
            self.assertIsNotNone(f)
            wb = load_workbook(f)
            ws = wb.active
            cells_values = []
            for row_number, row in enumerate(ws.rows):
                if row_number == row_nr:
                    for cell in row:
                        cells_values.append(cell.value)
            return cells_values

    def test_excel_single_robject_row_content(self):
        # set up db
        user, proj = self.default_set_up_for_robjects_pages()
        tag1 = Tag.objects.create(name="tag_1")
        tag2 = Tag.objects.create(name="tag_2")
        name1 = Name.objects.create(name="name_1")
        name2 = Name.objects.create(name="name_2")

        r = Robject.objects.create(
            author=user,
            project=proj,
            name="robject_1",
            create_by=user,
            modify_by=user,
            notes="robject_1_notes",
            ref_seq="robject_1_ref_seq",
            mod_seq="robject_1_mod_seq",
            description="robject_1_description",
            bibliography="robject_1_bibliography",
            ref_commercial="robject_1_ref_commercial",
            ref_clinical="robject_1_ref_clinical",
            ligand="robject_1_ligand",
            receptor="robject_1_receptor",
        )

        r.names.add(name1, name2)
        r.tags.add(tag1, tag2)
        r.save()

        response = self.client.get(
            f"/projects/{proj.name}/robjects/excel-raport/",
            {"robject_1": r.id}
        )

        cells_values = self.get_cells_values_from_excel_row(response, row_nr=1)

        # following order must be preserved:
        # 'id', 'project', 'author', 'name', 'create_by',
        # 'modify_by', 'notes', 'ref_seq', 'mod_seq',
        # 'description', 'bibliography', 'ref_commercial', 'ref_clinical',
        # 'ligand', 'receptor', 'tags', 'names'
        expected_values = [
            str(r.id), r.project.name, r.author.username, r.name,
            r.create_by.username, r.create_date.strftime("%Y-%m-%d %H:%M"),
            r.modify_by.username, r.modify_date.strftime("%Y-%m-%d %H:%M"),
            r.notes, r.ref_seq, r.mod_seq, r.description, r.bibliography,
            r.ref_commercial, r.ref_clinical, r.ligand, r.receptor,
            r.tags.all().all_as_string(), r.names.all().all_as_string()
        ]

        self.assertEqual(expected_values, cells_values)

    def test_excel_another_robject_row_content(self):
        """ This test is similar to previous except changed data in models
        """
        # set up db
        user, proj = self.default_set_up_for_robjects_pages()
        tag1 = Tag.objects.create(name="tag_3")
        tag2 = Tag.objects.create(name="tag_4")
        name1 = Name.objects.create(name="name_3")
        name2 = Name.objects.create(name="name_4")

        r = Robject.objects.create(
            author=user,
            project=proj,
            name="robject_2",
            create_by=user,
            modify_by=user,
            notes="robject_2_notes",
            ref_seq="robject_2_ref_seq",
            mod_seq="robject_2_mod_seq",
            description="robject_2_description",
            bibliography="robject_2_bibliography",
            ref_commercial="robject_2_ref_commercial",
            ref_clinical="robject_2_ref_clinical",
            ligand="robject_2_ligand",
            receptor="robject_2_receptor",
        )

        r.names.add(name1, name2)
        r.tags.add(tag1, tag2)
        r.save()

        response = self.client.get(
            f"/projects/{proj.name}/robjects/excel-raport/",
            {"robject_2": r.id}
        )

        cells_values = self.get_cells_values_from_excel_row(response, row_nr=1)

        # following order must be preserved:
        # 'id', 'project', 'author', 'name', 'create_by',
        # 'modify_by', 'notes', 'ref_seq', 'mod_seq',
        # 'description', 'bibliography', 'ref_commercial', 'ref_clinical',
        # 'ligand', 'receptor', 'tags', 'names'
        expected_values = [
            str(r.id), r.project.name, r.author.username, r.name,
            r.create_by.username, r.create_date.strftime("%Y-%m-%d %H:%M"),
            r.modify_by.username, r.modify_date.strftime("%Y-%m-%d %H:%M"),
            r.notes, r.ref_seq, r.mod_seq, r.description, r.bibliography,
            r.ref_commercial, r.ref_clinical, r.ligand, r.receptor,
            r.tags.all().all_as_string(), r.names.all().all_as_string()
        ]

        self.assertEqual(expected_values, cells_values)

    def test_excel_multiple_rows_content(self):
        # set up db
        user, proj = self.default_set_up_for_robjects_pages()
        tag1 = Tag.objects.create(name="tag_1")
        tag2 = Tag.objects.create(name="tag_2")
        name1 = Name.objects.create(name="name_1")
        name2 = Name.objects.create(name="name_2")

        r1 = Robject.objects.create(
            author=user,
            project=proj,
            name="robject_1",
            create_by=user,
            modify_by=user,
            notes="robject_1_notes",
            ref_seq="robject_1_ref_seq",
            mod_seq="robject_1_mod_seq",
            description="robject_1_description",
            bibliography="robject_1_bibliography",
            ref_commercial="robject_1_ref_commercial",
            ref_clinical="robject_1_ref_clinical",
            ligand="robject_1_ligand",
            receptor="robject_1_receptor",
        )

        r2 = Robject.objects.create(
            author=user,
            project=proj,
            name="robject_2",
            create_by=user,
            modify_by=user,
            notes="robject_2_notes",
            ref_seq="robject_2_ref_seq",
            mod_seq="robject_2_mod_seq",
            description="robject_2_description",
            bibliography="robject_2_bibliography",
            ref_commercial="robject_2_ref_commercial",
            ref_clinical="robject_2_ref_clinical",
            ligand="robject_2_ligand",
            receptor="robject_2_receptor",
        )

        r1.names.add(name1)
        r1.tags.add(tag1)
        r1.save()

        r2.names.add(name2)
        r2.tags.add(tag2)
        r2.save()

        response = self.client.get(
            f"/projects/{proj.name}/robjects/excel-raport/",
            {"robject_1": r1.id, "robject_2": r2.id}
        )

        row_1_cells_values = self.get_cells_values_from_excel_row(
            response, row_nr=1)

        # following order must be preserved:
        # 'id', 'project', 'author', 'name', 'create_by',
        # 'modify_by', 'notes', 'ref_seq', 'mod_seq',
        # 'description', 'bibliography', 'ref_commercial', 'ref_clinical',
        # 'ligand', 'receptor', 'tags', 'names'
        row_1_expected_values = [
            str(r1.id), r1.project.name, r1.author.username, r1.name,
            r1.create_by.username, r1.create_date.strftime("%Y-%m-%d %H:%M"),
            r1.modify_by.username, r1.modify_date.strftime("%Y-%m-%d %H:%M"),
            r1.notes, r1.ref_seq, r1.mod_seq, r1.description, r1.bibliography,
            r1.ref_commercial, r1.ref_clinical, r1.ligand, r1.receptor,
            r1.tags.all().all_as_string(), r1.names.all().all_as_string()
        ]

        self.assertEqual(row_1_expected_values, row_1_cells_values)

        row_2_cells_values = self.get_cells_values_from_excel_row(
            response, row_nr=2)

        # following order must be preserved:
        # 'id', 'project', 'author', 'name', 'create_by',
        # 'modify_by', 'notes', 'ref_seq', 'mod_seq',
        # 'description', 'bibliography', 'ref_commercial', 'ref_clinical',
        # 'ligand', 'receptor', 'tags', 'names'
        row_2_expected_values = [
            str(r2.id), r2.project.name, r2.author.username, r2.name,
            r2.create_by.username, r2.create_date.strftime("%Y-%m-%d %H:%M"),
            r2.modify_by.username, r2.modify_date.strftime("%Y-%m-%d %H:%M"),
            r2.notes, r2.ref_seq, r2.mod_seq, r2.description, r2.bibliography,
            r2.ref_commercial, r2.ref_clinical, r2.ligand, r2.receptor,
            r2.tags.all().all_as_string(), r2.names.all().all_as_string()
        ]

        self.assertEqual(row_2_expected_values, row_2_cells_values)

    def test_error_message_when_no_selection(self):
        user, proj = self.default_set_up_for_robjects_pages()
        response = self.client.get(f"/projects/{proj.name}/robjects/excel-raport/", follow=True)
        message = list(response.context['messages'])[0]
        self.assertEqual(
            message.message,
            "No robject selected!"
        )
        self.assertEqual(message.tags, "error")


class RobjectSamplesListTest(FunctionalTest):
    def test_view_returns_404_when_slug_not_match(self):
        self.not_matching_url_slug_helper(self.SAMPLE_LIST_URL)

    def test_anonymous_user_gets_samples_page(self):
        proj = Project.objects.create(name="PROJECT_1")
        Robject.objects.create(name="Robject_1", project=proj)
        response = self.client.get("/projects/PROJECT_1/robjects/1/samples/")
        self.assertEqual(response.status_code, 302)
        self.assertIn('/accounts/login/?next=', response.url)

    def test_render_template_on_get(self):
        user, proj = self.default_set_up_for_robjects_pages()
        robj = Robject.objects.create(name="rob")
        assign_perm("projects.can_visit_project", user, proj)
        samp = Sample(code='1a2b3c')
        response = self.client.get(f"/projects/{proj.name}/robjects/{robj.id}/samples/")
        self.assertTemplateUsed(response, "samples/samples_list.html")

    def test_view_get_list_of_samples_and_pass_it_to_context(self):
        user, proj = self.default_set_up_for_robjects_pages()
        assign_perm("projects.can_visit_project", user, proj)

        robj = Robject.objects.create(name='robject', project=proj)

        samp1 = Sample.objects.create(code='1a1a', robject=robj)
        samp2 = Sample.objects.create(code='2a2a', robject=robj)
        samp3 = Sample.objects.create(code='3a3a', robject=robj)
        response = self.client.get(f"/projects/{proj.name}/robjects/{robj.id}/samples/")

        self.assertIn(samp1, response.context["sample_list"])
        self.assertIn(samp2, response.context["sample_list"])
        self.assertIn(samp3, response.context["sample_list"])

    def test_context_data(self):
        user, proj = self.default_set_up_for_robjects_pages()
        assign_perm("projects.can_visit_project", user, proj)

        robj = Robject.objects.create(name='robject', project=proj)

        samp1 = Sample.objects.create(code='1a1a', robject=robj)
        response = self.client.get(f"/projects/{proj.name}/robjects/{robj.id}/samples/")
        self.assertEqual(proj, response.context['project'])


class RObjectsListViewTests(FunctionalTest):
    def test_view_returns_404_when_slug_not_match(self):
        self.not_matching_url_slug_helper(self.ROBJECT_LIST_URL)

    def test_anonymous_user_gets_robjects_page(self):
        Project.objects.create(name="PROJECT_1")
        response = self.client.get("/projects/PROJECT_1/robjects/")
        self.assertEqual(response.status_code, 302)

    def test_render_template_on_get(self):
        user, proj = self.default_set_up_for_robjects_pages()
        response = self.client.get(f"/projects/{proj.name}/robjects/")

        self.assertTemplateUsed(response, "projects/robjects_list.html")

    def test_view_create_list_of_robjects_and_pass_it_to_context(self):
        user, proj = self.default_set_up_for_robjects_pages()

        robj1 = Robject.objects.create(author=user, project=proj)
        robj2 = Robject.objects.create(author=user, project=proj)
        robj3 = Robject.objects.create(author=user, project=proj)
        response = self.client.get(f"/projects/{proj.name}/robjects/")

        self.assertIn(robj1, response.context["robject_list"])
        self.assertIn(robj2, response.context["robject_list"])


class SearchRobjectsViewTests(FunctionalTest):
    def test_view_returns_404_when_slug_not_match(self):
        self.not_matching_url_slug_helper(self.ROBJECT_SEARCH_URL)

    def test_view_renders_robjects_page_template(self):
        user, proj = self.default_set_up_for_robjects_pages()

        response = self.client.get(f"/projects/{proj.name}/robjects/search/",
                                   {"query": ""})
        self.assertTemplateUsed(response, "projects/robjects_list.html")

    def test_view_gets_valid_query_on_get__view_pass_qs_to_template(self):
        user, proj = self.default_set_up_for_robjects_pages()

        robject_1 = Robject.objects.create(name="robject_1", project=proj)
        robject_2 = Robject.objects.create(name="robject_2", project=proj)

        response = self.client.get(
            f"/projects/{proj.name}/robjects/search/", {"query": "robject_1"})
        queryset = Robject.objects.filter(name="robject_1")
        # comparison of two querysets
        self.assertQuerysetEqual(
            response.context["robject_list"], map(repr, queryset))

    def test_annonymous_user_has_no_access_to_search_view(self):
        proj = Project.objects.create(name="project_1")
        requested_url = f"/projects/{proj.name}/robjects/search/"
        resp = self.client.get(requested_url)
        self.assertRedirects(resp, reverse("login") + f"?next={requested_url}")

    def test_view_can_perform_search_basing_on_part_of_robject_name(self):
        user, proj = self.default_set_up_for_robjects_pages()

        robject_1 = Robject.objects.create(name="robject_1", project=proj)
        robject_2 = Robject.objects.create(name="robject_2", project=proj)

        response = self.client.get(
            f"/projects/{proj.name}/robjects/search/", {"query": "object_1"})  # part!

        queryset = Robject.objects.filter(name="robject_1")

        # comparison of two querysets
        self.assertQuerysetEqual(
            response.context["robject_list"], map(repr, queryset))

    def test_search_is_case_insensitive(self):
        user, proj = self.default_set_up_for_robjects_pages()

        robj = Robject.objects.create(name="RoBjEcT_1", project=proj)

        # lower case query
        resp = self.client.get(
            f"/projects/{proj.name}/robjects/search/", {"query": "robject_1"})

        self.assertEqual(list(resp.context["robject_list"]),
                         [robj])

        # upper case query
        resp = self.client.get(
            f"/projects/{proj.name}/robjects/search/", {"query": "ROBJECT_1"})

        self.assertEqual(list(resp.context["robject_list"]),
                         [robj])

    def test_view_pass_project_name_to_context(self):
        user, proj = self.default_set_up_for_robjects_pages()

        resp = self.client.get(f"/projects/{proj.name}/robjects/search/",
                               {"query": "robject_1"})

        self.assertEqual(proj.name, resp.context["project_name"])

    def create_sample_robject_and_send_query_to_search_view(self, project,
                                                            query,
                                                            **robject_kwargs):

        robj = Robject.objects.create(project=project, **robject_kwargs)
        resp = self.client.get(f"/projects/{project.name}/robjects/search/",
                               {"query": query})
        return robj, resp

    def create_sample_robject_search_for_it_and_confirm_results(
            self, query, robject_kwargs):
        user, proj = self.default_set_up_for_robjects_pages()

        robj, resp = self.create_sample_robject_and_send_query_to_search_view(
            project=proj, query=query, **robject_kwargs)

        self.assertIn(robj, resp.context["robject_list"])

    def test_search_include_full_author_username(self):
        robject_kwargs = {
            "author": User.objects.create_user(username="AUTHOR")
        }
        self.create_sample_robject_search_for_it_and_confirm_results(
            query="AUTHOR",
            robject_kwargs=robject_kwargs
        )

    def test_search_include_fragment_author_username(self):
        robject_kwargs = {
            "author": User.objects.create_user(username="AUTHOR")
        }
        self.create_sample_robject_search_for_it_and_confirm_results(
            query="AUTH",
            robject_kwargs=robject_kwargs
        )

    def test_search_include_case_insensitive_full_author_username(self):
        robject_kwargs = {
            "author": User.objects.create_user(username="AUTHOR")
        }
        self.create_sample_robject_search_for_it_and_confirm_results(
            query="aUtHoR",
            robject_kwargs=robject_kwargs
        )

    def test_empty_query_will_display_all_robjects(self):
        user, proj = self.default_set_up_for_robjects_pages()

        robj_1 = Robject.objects.create(project=proj)
        robj_2 = Robject.objects.create(project=proj)

        resp = self.client.get(
            f"/projects/{proj.name}/robjects/search/", {"query": ""})

        all_robjects = Robject.objects.filter(project=proj)

        self.assertEqual(
            list(resp.context["robject_list"]),
            list(all_robjects)
        )

    def test_search_include_robjects_from_given_project(self):
        user, proj = self.default_set_up_for_robjects_pages()

        other_proj = Project.objects.create(name="other_proj")
        robj = Robject.objects.create(project=other_proj, name="robj")

        resp = self.client.get(
            f"/projects/{proj.name}/robjects/search/",
            {"query": f"{robj.name}"})

        self.assertNotIn(
            robj,
            list(resp.context["robject_list"])
        )


class RobjectCreateViewTestCase(FunctionalTest):
    def get_robject_create_url(self, proj):
        return reverse("projects:robjects:robject_create", kwargs={"project_name": proj.name})

    def get_form_from_context(self):
        user, proj = self.default_set_up_for_robjects_pages()
        assign_perm("projects.can_modify_project", user, proj)
        response = self.client.get(
            reverse("projects:robjects:robject_create", kwargs={"project_name": proj.name}))
        form = response.context["form"]

        return form

    def test_view_returns_404_when_slug_not_match(self):
        self.not_matching_url_slug_helper(self.ROBJECT_CREATE_URL)

    def test_renders_template(self):
        user, proj = self.default_set_up_for_robjects_pages()
        assign_perm("projects.can_modify_project", user, proj)
        response = self.client.get(
            reverse("projects:robjects:robject_create", args=(proj.name,)))
        self.assertTemplateUsed(
            response, template_name="robjects/robject_create.html")

    def test_renders_form_in_context(self):
        user, proj = self.default_set_up_for_robjects_pages()
        assign_perm("projects.can_modify_project", user, proj)
        response = self.client.get(
            reverse("projects:robjects:robject_create", args=(proj.name,)))

        self.assertEqual(response.status_code, 200)
        self.assertIn("form", response.context)

    def test_model_class_is_Robject_in_form(self):
        form = self.get_form_from_context()
        self.assertEqual(form._meta.model, Robject)

    def test_widget_instance_of_names_field_in_form(self):
        form = self.get_form_from_context()
        self.assertIsInstance(
            form.base_fields["names"].widget, AddAnotherWidgetWrapper)

    def test_names_widget_arguments_in_form(self):
        form = self.get_form_from_context()
        widget = form.base_fields["names"].widget.widget
        add_related_url = form.base_fields["names"].widget.add_related_url
        self.assertIsInstance(widget, forms.SelectMultiple)
        self.assertEqual(add_related_url, reverse(
            "projects:robjects:names_create", args=(Project.objects.last().name,)))

    def test_widget_instance_of_tags_field_in_form(self):
        form = self.get_form_from_context()
        self.assertIsInstance(
            form.base_fields["tags"].widget, AddAnotherWidgetWrapper)

    def test_tags_widget_arguments_in_form(self):
        form = self.get_form_from_context()
        widget = form.base_fields["tags"].widget.widget
        add_related_url = form.base_fields["tags"].widget.add_related_url
        self.assertIsInstance(widget, forms.SelectMultiple)
        self.assertEqual(add_related_url, reverse(
            "projects:robjects:tags_create", args=(Project.objects.last().name,)))

    def test_view_redirects_to_robject_list_page_on_post(self):
        user, proj = self.default_set_up_for_robjects_pages()
        assign_perm("projects.can_modify_project", user, proj)
        response = self.client.post(
            self.get_robject_create_url(proj),
            data={"name": "whatever"}
        )
        self.assertRedirects(response, reverse(
            "projects:robjects:robjects_list", args=(proj.name,)))

    def test_name_field_is_required(self):
        user, proj = self.default_set_up_for_robjects_pages()
        assign_perm("projects.can_modify_project", user, proj)
        response = self.client.post(
            self.get_robject_create_url(proj),
            data={"hello": "world"}  # request.POST pass, form.is_valid() fails
        )

        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            self.get_robject_create_url(proj),
            data={"name": "whatever"}
        )

        self.assertEqual(response.status_code, 302)

    def test_annonymous_user_is_redirect_to_login_page_on_get(self):
        proj = Project.objects.create(name="proj")
        response = self.client.get(self.get_robject_create_url(proj))

        self.assertRedirects(
            response,
            settings.LOGIN_URL + "?next=" + self.get_robject_create_url(proj)
        )

    def test_annonymous_user_is_redirect_to_login_page_on_post(self):
        proj = Project.objects.create(name="proj")
        response = self.client.post(self.get_robject_create_url(proj), data={})

        self.assertRedirects(
            response,
            settings.LOGIN_URL + "?next=" + self.get_robject_create_url(proj)
        )

    def test_user_without_project_mod_permission_gets_403_on_get(self):
        user, proj = self.default_set_up_for_robjects_pages()
        response = self.client.get(self.get_robject_create_url(proj))

        self.assertEqual(response.status_code, 403)

    def test_view_removes_all_robject_less_names_on_get(self):
        Name.objects.create(name="name_1")
        Name.objects.create(name="name_2")
        self.assertEqual(Name.objects.filter(robjects=None).count(), 2)
        user, proj = self.default_set_up_for_robjects_pages()
        assign_perm("projects.can_modify_project", user, proj)
        self.client.get(self.get_robject_create_url(proj))
        self.assertEqual(Name.objects.filter(robjects=None).count(), 0)

    def test_rendered_form_has_no_create_by_field(self):
        form = self.get_form_from_context()
        self.assertNotIn("create_by", form.fields)

    def test_rendered_form_has_no_modify_by_field(self):
        form = self.get_form_from_context()
        self.assertNotIn("modify_by", form.fields)

    def test_rendered_form_has_no_create_date_field(self):
        form = self.get_form_from_context()
        self.assertNotIn("create_date", form.fields)

    def test_rendered_form_has_no_modify_date_field(self):
        form = self.get_form_from_context()
        self.assertNotIn("modify_date", form.fields)

    def test_view_assign_create_by_to_new_robject(self):
        user, proj = self.default_set_up_for_robjects_pages()
        assign_perm("projects.can_modify_project", user, proj)
        response = self.client.post(
            self.get_robject_create_url(proj), {"name": "test"})
        r = Robject.objects.last()
        self.assertEqual(r.create_by, user)

    def test_view_assign_modify_by_to_new_robject(self):
        user, proj = self.default_set_up_for_robjects_pages()
        assign_perm("projects.can_modify_project", user, proj)
        response = self.client.post(
            self.get_robject_create_url(proj), {"name": "test"})
        r = Robject.objects.last()
        self.assertEqual(r.modify_by, user)

    def test_project_field_from_context_form_is_hidden(self):
        form = self.get_form_from_context()
        self.assertTrue(form.fields["project"].widget.is_hidden)

    def test_project_field_from_context_form_initial(self):
        form = self.get_form_from_context()
        self.assertEqual(form.initial, {"project": Project.objects.first()})


class NameCreateViewTestCase(FunctionalTest):
    def get_names_create_url(self, proj):
        url = reverse("projects:robjects:names_create",
                      kwargs={"project_name": proj.name})
        return url

    def get_robject_create_url(self, proj):
        return reverse("projects:robjects:robject_create", args=(proj.name,))

    def test_view_parents(self):
        self.assertEqual(NameCreateView.__bases__,
                         (CreatePopupMixin, generic.CreateView))

    def test_view_model_attr(self):
        self.assertEqual(NameCreateView.model, Name)

    def test_view_fields_attr(self):
        self.assertEqual(NameCreateView.fields, "__all__")

    def test_render_template(self):
        proj = Project.objects.create(name="proj_1")
        response = self.client.get(
            self.get_names_create_url(proj),
            HTTP_REFERER=self.get_robject_create_url(proj))
        self.assertTemplateUsed(response, "robjects/names_create.html")

    def test_view_return_400_when_requested_using_url(self):
        proj = Project.objects.create(name="proj_1")

        # Note: when form is open in popup window, request contains HTTP_REFERER
        # key in request.META dictionary. Otherwise, key doesn't exists.
        # HTTP_REFERER holds url to page from popup is open.

        # Test view gets request object without "HTTP_REFERER" key.
        response = self.client.get(self.get_names_create_url(proj))
        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            "<h1>Error 400</h1><p>Form available from robject form only</p>",
            response.content.decode())

        # Test view gets request with META["HTTP_REFERER"].
        response = self.client.get(
            self.get_names_create_url(proj),
            HTTP_REFERER=self.get_robject_create_url(proj))
        self.assertEqual(response.status_code, 200)


class TagCreateViewTestCase(FunctionalTest):
    def get_tags_create_url(self, proj):
        url = reverse("projects:robjects:tags_create",
                      kwargs={"project_name": proj.name})
        return url

    def get_robject_create_url(self, proj):
        return reverse("projects:robjects:robject_create", args=(proj.name,))

    def test_view_parents(self):
        self.assertEqual(TagCreateView.__bases__,
                         (CreatePopupMixin, generic.CreateView))

    def test_view_model_attr(self):
        self.assertEqual(TagCreateView.model, Tag)

    def test_render_template(self):
        proj = Project.objects.create(name="proj_1")
        response = self.client.get(reverse("projects:robjects:tags_create", kwargs={
                                   "project_name": proj.name}), HTTP_REFERER=self.get_robject_create_url(proj))
        self.assertTemplateUsed(response, "robjects/tags_create.html")

    def test_view_renders_form_without_project_field(self):
        proj = Project.objects.create(name="proj_1")
        response = self.client.get(reverse("projects:robjects:tags_create", kwargs={
                                   "project_name": proj.name}), HTTP_REFERER=self.get_robject_create_url(proj))
        self.assertNotIn("project", response.context["form"].fields)

    def test_project_is_assigned_automatically_in_view(self):
        proj = Project.objects.create(name="proj_1")
        self.client.post(
            reverse("projects:robjects:tags_create", kwargs={"project_name": proj.name}), data={"name": "tag_name"},
            HTTP_REFERER=self.get_robject_create_url(proj))
        t = Tag.objects.last()
        self.assertEqual(t.name, "tag_name")
        self.assertEqual(t.project, proj)

    def test_view_return_400_when_requested_using_url(self):
        proj = Project.objects.create(name="proj_1")

        # Note: when form is open in popup window, request contains HTTP_REFERER
        # key in request.META dictionary. Otherwise, key doesn't exists.
        # HTTP_REFERER holds url to page from popup is open.

        # Test view gets request object without "HTTP_REFERER" key.
        response = self.client.get(self.get_tags_create_url(proj))
        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            "<h1>Error 400</h1><p>Form available from robject form only</p>",
            response.content.decode())

        # Test view gets request with META["HTTP_REFERER"].
        response = self.client.get(
            self.get_tags_create_url(proj),
            HTTP_REFERER=self.get_robject_create_url(proj))
        self.assertEqual(response.status_code, 200)


class RobjectDeleteTestCase(FunctionalTest):
    def test_view_returns_404_when_slug_not_match(self):
        self.not_matching_url_slug_helper(self.ROBJECT_DELETE_URL)

    def default_set_up_for_robject_delete(self):
        user, proj = self.default_set_up_for_robjects_pages()
        assign_perm("can_modify_project", user, proj)
        return proj

    def test_annonymous_user_is_redirect_to_login_page(self):
        self.annonymous_testing_helper(self.ROBJECT_DELETE_URL)

    def test_view_refuse_access_to_users_without_both_permissions(self):
        user = self.default_set_up_for_projects_pages()
        proj = Project.objects.create(name="project_1")
        response = self.client.get(self.ROBJECT_DELETE_URL)
        self.assertEqual(response.status_code, 403)
        self.assertEqual(f"<h1>User doesn't have permission: can visit project</h1>",
                         response.content.decode("utf-8"))

    def test_view_refuse_access_to_users_without_project_modify_permission(self):
        self.permission_testing_helper(
            self.ROBJECT_DELETE_URL,
            preassigned_perms=["projects.can_visit_project"],
            error_message="User doesn't have permission: can modify project")

    def test_view_refuse_access_to_users_without_project_visit_permission(self):
        user = self.default_set_up_for_projects_pages()
        proj = Project.objects.create(name="project_1")
        assign_perm("projects.can_modify_project", user, proj)
        response = self.client.get(self.ROBJECT_DELETE_URL)
        self.assertEqual(response.status_code, 403)
        self.assertEqual(f"<h1>User doesn't have permission: can visit project</h1>",
                         response.content.decode("utf-8"))

    def test_view_renders_delete_confirmattion_template(self):
        self.default_set_up_for_robject_delete()
        response = self.client.get(self.ROBJECT_DELETE_URL)
        self.assertTemplateUsed(
            response, "robjects/robject_confirm_delete.html")

    def test_view_redirects_on_post(self):
        self.default_set_up_for_robject_delete()
        response = self.client.post(self.ROBJECT_DELETE_URL)
        self.assertRedirects(response, self.ROBJECT_LIST_URL)

    def test_view_add_delete_robjects_to_context(self):
        proj = self.default_set_up_for_robject_delete()
        robj_1 = Robject.objects.create(name="robject_1", project=proj)
        robj_2 = Robject.objects.create(name="robject_2", project=proj)
        response = self.client.get(self.ROBJECT_DELETE_URL, {
            robj_1.name: robj_1.id,
            robj_2.name: robj_2.id
        })
        robjects_context = response.context["robjects"]
        self.assertEqual(len(robjects_context), 2)
        self.assertIn(robj_1, robjects_context)
        self.assertIn(robj_2, robjects_context)


class RobjectEditView(FunctionalTest):
    def test_view_returns_404_when_slug_not_match(self):
        self.not_matching_url_slug_helper(self.ROBJECT_EDIT_URL)

    def test_view_render_bounded_form(self):
        user, proj = self.default_set_up_for_robjects_pages()
        assign_perm("can_modify_project", user, proj)
        r = Robject.objects.create(project=proj, name="ROBJECT_NAME")
        response = self.client.get(self.ROBJECT_EDIT_URL)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.context["form"].initial["name"], 'ROBJECT_NAME')

    def test_view_updates_modify_by(self):
        user, proj = self.default_set_up_for_robjects_pages()
        r = Robject.objects.create(
            project=proj, name="ROBJECT_NAME", create_by=user, modify_by=user)
        new_user = User.objects.create_user(
            username="new_user", password="new_password")
        assign_perm("can_visit_project", new_user, proj)
        assign_perm("can_modify_project", new_user, proj)
        self.client.login(username="new_user", password="new_password")
        response = self.client.post(
            self.ROBJECT_EDIT_URL, {"name": "new_name"})
        self.assertEqual(Robject.objects.last().create_by.username, "USERNAME")
        self.assertEqual(Robject.objects.last().modify_by.username, "new_user")

    def test_view_redirects_on_post(self):
        user, proj = self.default_set_up_for_robjects_pages()
        assign_perm("can_modify_project", user, proj)
        r = Robject.objects.create(
            project=proj, name="ROBJECT_NAME", create_by=user, modify_by=user)
        response = self.client.post(
            self.ROBJECT_EDIT_URL, {"name": "new_name"})

        self.assertRedirects(response, reverse(
            "projects:robjects:robjects_list", kwargs={"project_name": proj.name}))


class RobjectsPdfTestCase(FunctionalTest):
    def test_view_returns_404_when_slug_not_match(self):
        self.not_matching_url_slug_helper(self.ROBJECT_PDF_URL)

    def test_anonymous_user_gets_robject_raport_page(self):
        proj = Project.objects.create(name="PROJECT_1")
        Robject.objects.create(name="Robject_1", project=proj)
        response = self.client.get(
            f"/projects/{proj.name}/robjects/PDF-raport/")
        self.assertEqual(response.status_code, 302)
        self.assertIn('/accounts/login/?next=', response.url)

    def test_render_template_on_get(self):
        user, proj = self.default_set_up_for_robjects_pages()
        robj = Robject.objects.create(name="rob", project=proj)
        assign_perm("projects.can_visit_project", user, proj)
        response = self.client.get(
            f"/projects/{proj.name}/robjects/PDF-raport/", {robj.name: robj.id})
        self.assertTemplateUsed(response, "robjects/robject_raport_pdf.html")

    def test_user_generates_pdf(self):
        # logged user goes to biodb to export a excel file
        user, proj = self.default_set_up_for_robjects_pages()
        robj = Robject.objects.create(
            author=user, project=proj, name="robject_1")

        response = self.client.get(
            f"/projects/{proj.name}/robjects/PDF-raport/", {robj.name: robj.id})
        # assert attachment name as robject_raport.pdf
        self.assertEqual(response.get('Content-Disposition'),
                         'filename="raport.pdf"')
        # check if pdf_file is not empty
        pdf_file = BytesIO(response.content)
        self.assertIsNotNone(pdf_file)
        # open and read pdf file
        read_pdf = PyPDF2.PdfFileReader(pdf_file, strict=False)
        number_of_pages = read_pdf.getNumPages()
        page = read_pdf.getPage(0)
        page_content = page.extractText()
        # chcek if robjects name is in page content
        self.assertIn('robject_1', page_content)
        # if status code of requeszt is 200 -
        # The request was successfully received,
        # understood, and accepted
        self.assertEqual(response.status_code, 200)

    def test_robjects_pdf_view_for_multiple_robjects(self):
        # CREATE SAMPLE PROJECT AND USER.
        # ASSIGNE PERMISION TO PROJECT.
        user, proj = self.default_set_up_for_robjects_pages()
        # CREATE SAMPLE ROBJECTS AND ADD IT TO PROJECT.
        robj1 = Robject.objects.create(
            author=user, project=proj, name="robject_1")
        robj2 = Robject.objects.create(
            author=user, project=proj, name="robject_2")
        # User enters project robjects pdf raport page.
        response = self.client.get(
            f"/projects/{proj.name}/robjects/PDF-raport/", {
                robj1.name: robj1.id, robj2.name: robj2.id})
        # User seas name of file.
        self.assertEqual(response.get('Content-Disposition'),
                         'filename="raport.pdf"')
        # He checks if file is not empty.
        pdf_file = BytesIO(response.content)
        self.assertIsNotNone(pdf_file)
        # He opens and reads file.
        read_pdf = PyPDF2.PdfFileReader(pdf_file, strict=False)
        # He count numbers of pages.
        number_of_pages = read_pdf.getNumPages()
        # He seas two pages, one for every robjects.
        self.assertEqual(number_of_pages, 2)
        # He seas first page of pdf.
        page = read_pdf.getPage(0)
        page_content = page.extractText()
        # He seas that first page is for robject_1 model.
        self.assertIn('robject_1', page_content)
        # He checks second page for robject_2.
        page = read_pdf.getPage(1)
        page_content = page.extractText()
        self.assertIn('robject_2', page_content)

        # if status code of requeszt is 200 -
        # The request was successfully received,
        # understood, and accepted
        self.assertEqual(response.status_code, 200)


class RobjectHistoryViewTest(FunctionalTest):
    def test_view_returns_404_when_slug_not_match(self):
        self.not_matching_url_slug_helper(self.ROBJECT_HISTORY_URL)

    def test_anonymous_user_visit_page(self):
        proj = Project.objects.create(name="Project_1")
        robject = Robject.objects.create(name="Robject_1", project=proj)
        response = self.client.get("/projects/Project_1/robjects/1/history/")
        self.assertEqual(response.status_code, 302)
        self.assertIn(
            f'/accounts/login/?next=/projects/{proj.name}/robjects/{robject.pk}/history/',
            response.url)

    def test_view_permission_is_required_to_visit_page(self):
        self.permission_testing_helper(
            self.ROBJECT_HISTORY_URL, "User doesn't have permission: can visit project")

    def test_logged_user_canrender_template_on_get(self):
        user, proj = self.default_set_up_for_robjects_pages()
        robject = Robject.objects.create(name="Robject_1", project=proj)
        response = self.client.get(
            f"/projects/{proj.name}/robjects/{robject.pk}/history/")
        self.assertTemplateUsed(response, "robjects/robject_history.html")

    def test_variables_in_context(self):
        # set default data for user, project and permision to visit project
        user, proj = self.default_set_up_for_robjects_pages()
        # create robject
        robject = Robject.objects.create(name="Robject_1", project=proj)
        # render the template
        response = self.client.get(
            f"/projects/{proj.name}/robjects/{robject.pk}/history/")
        # check if versions are in context
        self.assertTrue("versions" in response.context)
        versions = response.context["versions"]
        # check that versions are the list type,
        # by defult created objec should have only creation history
        self.assertIsInstance(versions, list)
        self.assertEqual(len(versions), 1)
        # check the type of objects in list
        self.assertIsInstance(response.context["versions"][0], CustomHistory)
        # chcec that object is in context
        self.assertTrue("object" in response.context)
        # check that an context object is our robject
        context_object = response.context["object"]
        self.assertEqual(context_object, robject)

    def test_versions_for_edited_robject(self):
        user, proj = self.default_set_up_for_robjects_pages()
        robject = Robject.objects.create(name="Robject_1", project=proj)
        robject.name = "newname"
        robject.save()
        response = self.client.get(
            f"/projects/{proj.name}/robjects/{robject.pk}/history/")
        self.assertTrue("versions" in response.context)
        versions = response.context["versions"]
        self.assertIsInstance(versions, list)
        self.assertEqual(len(versions), 2)
        # get the versions from the list to check them
        version_created, version_changed = response.context["versions"]
        self.assertIsInstance(version_created, CustomHistory)
        self.assertIsInstance(version_changed, CustomHistory)
        # check the attributes of created version
        self.assertEqual(version_created.version_id, 1)
        self.assertEqual(version_created.modify_type, 'Created')
        self.assertCountEqual(version_created.exclude, ["id", "create_date",
                                                        "modify_date"])
        # check differ fields
        vcreated_diff_objects = version_created.get_diff_objects()
        self.assertEqual(vcreated_diff_objects, [])
        # check the attributes of changed version
        self.assertEqual(version_changed.version_id, 2)
        self.assertEqual(version_changed.modify_type, 'Changed')
        self.assertCountEqual(version_changed.exclude, ["id", "create_date",
                                                        "modify_date"])
        # check differ fields
        vchanged_diff_objects = version_changed.get_diff_objects()
        self.assertEqual(len(vchanged_diff_objects), 1)
        # check the diff field values
        diff_object = vchanged_diff_objects[0]
        self.assertEqual(diff_object.field, "name")
        self.assertEqual(diff_object.old_value, "Robject_1")
        self.assertEqual(diff_object.new_value, "newname")
